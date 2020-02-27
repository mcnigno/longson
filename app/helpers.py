from .models import Janus, Doc_list, Pdb, Category, SourceFiles, Sourcetype, Mscode
from app import db
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, PatternFill, Color, Side, Alignment, NamedStyle
from collections import OrderedDict
from flask import flash
from flask_appbuilder.models.sqla.interface import SQLAInterface
from config import UPLOAD_FOLDER
from .full_mdi import full_mdi, full_mdi_last_rev, fulmdi2, fulmdi3

from app import rq


from datetime import timedelta

janus_not_in_document_list = []
pdb_not_in_document_list = []

def janus_upload_from_txt(source):
    session = db.session
    janus_file = open(UPLOAD_FOLDER + source)
    row_number = 0
    count_janus = 0
    doc_list = [x.client_reference for x in session.query(Doc_list).all()]
    try:
        for line in janus_file:
            
            #Skip in Headers at line 1
            row_number += 1
            
            
            if row_number == 1:
                continue
            
            # Split Row by |
            row = str(line).split('|')
            
            # Create Janus row object
            #print('line_number:', row[0])
            janus = Janus(
        
                linenumber=row[0],
                cat=row[1],
                doc_reference=row[2],
                weight=row[9],
                title=row[11],
                milestone_chain=row[12],
                dbs=row[13],
                wbs=row[14],
                fcr_one=row[15],
                fcr_two=row[16],
                fcr_three=row[17],
                mscode=row[19],
                cumulative=row[20],
                obs=row[23],
                initial_plan_date=date_parse(row[24]),
                revised_plan_date=date_parse(row[25]),
                forecast_date=date_parse(row[26]),
                actual_date=date_parse(row[27])
            )
        
            janus.created_by_fk = '1'
            janus.client_reference_id = row[8]

            # Check if the janus doc is or not in document list
            
            if row[8] in doc_list:
                session.add(janus)
                count_janus += 1
            else:
                janus.ex_client_reference_id = row[8]
                janus.note = janus.client_reference_id + ' | No reference in Document List.'
                janus.client_reference_id = None
                session.add(janus)
                janus_not_in_document_list.append(janus)

        session.commit()
        return str(count_janus) + ' Janus Updated'
    except:
        return 'Janus FAIL: check your source file'

def janus_upload(source):
    session = db.session
    janus_file = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True, read_only=True)
    janus_sheet = janus_file.active
    #row_number = 0
    count_janus = 0
    # Set the document list with all janus document
    #documents = janus_update_document_list(source)
    
    print('******    ******   Document list updated')
    doclist_query = session.query(Doc_list).all()
    doc_list = [x.client_reference for x in doclist_query]
    
    print('Doc list Ready')
    janus_row_fail = 1
    try:
        for row in janus_sheet.iter_rows(min_row=2):
            janus_row_fail += 1
            print('janus___rowww            ****************',
                    row[8].value,
                    row[0].value,
                    row[1].value,
                    row[2].value,
                    )
            if row[32].value != 'NOT APPLICABLE':
                print('BEFORE PLANNED DATE')
                print(type(row[31].value), row[31].value == '#N/A')
                planned_date = row[31].value
                if row[31].value == '#N/A':
                    print('setting planning date')
                    planned_date = None
                    #planned_date = date_parse(row[32].value)
                print('AFTER PLANNED DATE', planned_date)
                
                
                janus = Janus(
                    linenumber=row[0].value,
                    cat=row[2].value,
                    doc_reference=row[1].value,
                    weight=row[9].value,
                    title=row[11].value,
                    milestone_chain=row[12].value,
                    dbs=row[13].value,
                    wbs=row[14].value,
                    fcr_one=row[15].value,
                    fcr_two=row[16].value,
                    fcr_three=row[17].value,
                    fcr_four=row[18].value,
                    mscode=row[19].value,
                    cumulative=row[20].value,
                    obs=row[23].value,
                    initial_plan_date=row[24].value, 
                    
                    forecast_date=row[26].value,
                    actual_date=row[27].value,
                    planned_date=planned_date,
                    pdb_issue=row[32].value,
                    revised_plan_date=row[33].value
                )  
                print('janus___rowww    NOT APPLICABLE        ****************')
                janus.created_by_fk = '1'
                
                janus.client_reference_id = row[8].value 
                
                
                # Check if the janus doc is or not in document list 
                try:
                    if row[8].value in doc_list:
                        
                        session.add(janus)
                        session.commit()
                        count_janus += 1
                    else:
                        
                        janus.ex_client_reference = row[8].value
                        janus.note = str(janus.client_reference_id) + ' | No reference in Document List.'
                        janus.client_reference_id = None
                        
                        session.add(janus)
                        session.commit()
                        #janus_not_in_document_list.append(janus)
                
                        #session.commit()
                        #return str(count_janus) + ' Janus Updated'
                        session.remove()
                        ft = session.query(Sourcetype).filter(Sourcetype.source_type == 'Janus').first()
                        fs = session.query(SourceFiles).filter(SourceFiles.source_type_id == ft.id).first() 
                        fs.description = 'READY ->' + str(janus_row_fail)
                        fs.changed_by_fk = '1'    
                        session.commit()
                except:
                    session.remove()
                    ft = session.query(Sourcetype).filter(Sourcetype.source_type == 'Janus').first()
                    fs = session.query(SourceFiles).filter(SourceFiles.source_type_id == ft.id).first() 
                    fs.description = 'FAIL on row ' + str(janus_row_fail)+ ' - check your file @ -> ' + row[1].value
                    fs.changed_by_fk = '1'    
                    session.commit()
                    return 'Janus FAIL: check your source file.'
                
        
    except:
        
        return 'Janus FAIL: check your source file'

def janus_update_document_list(source):
    
    session = db.session
    janus_file = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True)
    janus_sheet = janus_file.active
    #row_number = 0
    count_janus = 0
    doclist_query = session.query(Doc_list).all()
    doc_list = [x.client_reference for x in doclist_query]
    print('Doc list Ready')
    
    for row in janus_sheet.iter_rows(min_row=2):
        print('new document from janus___rowww            ****************',
                row[8].value,
                row[0].value,
                row[1].value,
                row[2].value,
                )
        if row[8].value not in doc_list and row[32].value != 'NOT APPLICABLE':
            
            document = Doc_list(
                
                doc_reference=row[1].value,
                weight=row[9].value,
                title=row[11].value,
                client_reference=row[8].value,
                cat=row[13].value[6:8],
                
                org=row[36].value
            )  
            
            document.created_by_fk = '1'
            
            session.add(document)
            print('Document created',document.client_reference )
            
                

    session.commit()
    return 'done'

@rq.job('low', timeout=3600)        
def janus_update():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    janusdel = session.query(Janus).delete()
    print('deleted from JANUS', janusdel)
    janus = janus_upload(files['Janus'])
    
    return janus

#janus_update()    
# 
def doclist_update_from_janus():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    
    source = files['Janus']
    janus_file = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True)
    janus_sheet = janus_file.active
    
    
    print('Doc list Ready')
    
    for row in janus_sheet.iter_rows(min_row=2):
        doclist_query = session.query(Doc_list).filter(
            Doc_list.client_reference==row[8].value).first()
        
        print('new document from janus___rowww            ****************',
                row[8].value,
                row[0].value,
                row[1].value,
                row[2].value,
                )
        if doclist_query is None and row[32].value != 'NOT APPLICABLE':
            print('CREATING A NEW DOCUMENT ** *** *** *** ***')
            
            document = Doc_list(
                
                doc_reference=row[1].value,
                weight=row[9].value,
                title=row[11].value,
                client_reference=row[8].value,
                cat=row[13].value[6:8],
                
                org=row[36].value
            )  
            
            
            
            session.add(document)
            print('Document created',document.client_reference )
            
        session.commit()
     
#doclist_update_from_janus()

def date_parse(date):
    try:
        print('Date is Type:', type(date), date)
        if isinstance(date, str): return None
        if isinstance(date, datetime): return date
        if str(date) == '#N/A' or str(date) == '#N/D': return None
        if date == 'NOT APPLICABLE':
            print(date) 
            return None
        if date == '' or date is None: return None

        #return datetime.strptime(date, '%d-%b-%y')
    except:
        print('something Wrong with DATE PARSE')
        return None

def document_list_upload(source):
    print('document_list_upload***')
    session = db.session

    doclist = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True, read_only=True)
    doclist_ws = doclist.active
    # count_id = 0
    count_doc = 0
    # wrong_ref = session.query(Doc_list).filter(Doc_list.client_reference == 'wrong_client_ref').first()
    # empty_client_references = session.query(Doc_list).filter(Doc_list.client_reference == 'EmptyClientReferences').first()            
    row_fail = 1
    try:    
        for row in doclist_ws.iter_rows(min_row=2):
            row_fail += 1
            doc = session.query(Doc_list).filter(
                Doc_list.client_reference == row[2].value
            ).first()

            if row[2].value and doc is None:
                doclist_row = Doc_list(
                    cat=row[0].value,
                    title=row[3].value,
                    org=row[5].value,
                    cat_class=row[6].value,
                    class_two=row[7].value,
                    weight=row[8].value,
                    doc_reference=row[12].value
                )
                doclist_row.created_by_fk = '1'
                doclist_row.client_reference = row[2].value
                count_doc += 1
                print('Document',row[2].value)
                session.add(doclist_row)
            
            if row[2].value and doc:
                doc.cat=row[0].value
                doc.title=row[3].value
                doc.org=row[5].value
                doc.cat_class=row[6].value
                doc.class_two=row[7].value
                doc.weight=row[8].value
                doc.doc_reference=row[12].value
                doc.changed_by_fk = '1'
                count_doc += 1
            session.commit()

        ft = session.query(Sourcetype).filter(Sourcetype.source_type == 'Document List').first()
        fs = session.query(SourceFiles).filter(SourceFiles.source_type_id == ft.id).first() 
        fs.description = 'READY'
        fs.changed_by_fk = '1'   
        print('Document in DC processed', count_doc)
        
        session.commit()
        return str(count_doc) + ' Document List updated!'
    except:
        session.remove()
        ft = session.query(Sourcetype).filter(Sourcetype.source_type == 'Document List').first()
        fs = session.query(SourceFiles).filter(SourceFiles.source_type_id == ft.id).first() 
        fs.description = 'FAIL on row ' + str(row_fail)+ ' - check your doc -> ' + row[2].value
        fs.changed_by_fk = '1'    
        session.commit()
        return 'Document List FAIL: check your source file.'

from rq import Worker, Queue, Connection

@rq.job('low', timeout=3600)
def document_list_update():
    '''
    with Connection():
        low_worker = Worker(['low'])
        low_worker.work()
    '''
    session = db.session  
    sf = SourceFiles
    files = dict([(str(x.source_type),x.file_source) for x in session.query(sf).all()])

    doc = document_list_upload(files['Document List'])
    session.close()

#document_list_update()
 
def pdb_list_upload(source):
    session = db.session
    pdblist = openpyxl.load_workbook(UPLOAD_FOLDER + source)
    pdblist_ws = pdblist.active
    session.query(Pdb).delete()
 
    doc_list = [x.client_reference for x in session.query(Doc_list).all()]
    
    count_pdb = 0
    try:           
        for row in pdblist_ws.iter_rows(min_row=2):
             if row[5].value is not None:
                print('        ----------------- HERE --------------',row[0].value,row[5].value)
                
                pdb = Pdb(
                    doc_reference=row[0].value,
                    title=row[1].value,
                    revision_number=row[2].value,
                    revision_date=date_parse(row[3].value),
                    document_revision_object=row[4].value,

                    discipline=row[6].value,
                    transmittal_date=date_parse(row[7].value),
                    transmittal_reference=row[8].value,
                    specific_transmittal_number=row[9].value,
                    required_action=row[10].value,
                    response_due_date=date_parse(row[11].value),
                    actual_response_date=date_parse(row[12].value),
                    document_status=row[13].value,
                    client_transmittal_ref_number=row[14].value,
                    remarks=row[15].value,
                )
                pdb.created_by_fk = '1'
                pdb.client_reference_id = row[5].value
                

                
                # Check if the pdb doc is or not in document list
                
                if row[5].value in doc_list:
                    session.add(pdb)
                    count_pdb += 1
                else:
                    pdb.ex_client_reference = pdb.client_reference_id
                    pdb.note = pdb.client_reference_id + ' | No reference in Document List.'
                    pdb.client_reference_id = None
                    
                    session.add(pdb)
                    pdb_not_in_document_list.append(pdb)
                

        session.commit()
        return str(count_pdb) + ' PDB updated!'
    except:
        return 'PDB FAIL: check your source file.'  

#pdb_list_upload()

def pdb_replace():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    pdbdel = session.query(Pdb).delete()
    print('deleted from PDB', pdbdel)
    
    pdb = pdb_list_upload(files['PDB'])
    print(pdb)

    return pdb

#pdb_update()
# Open the Category Code List 
def category_upload(source):
    wcb = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True, read_only=True)
    sheet_list = wcb.sheetnames
    
    session = db.session
    count_cat = 0
    
    for name in sheet_list[6:]:
        print(' - - - - - - - - Category - - - - ')
        print(name)
        wcs = wcb[name]
        
        for row in wcs.iter_rows(min_row=9):
            #print(' - - - - - - - - Row - - - - ',row[1].value,row[12].value[0:1])

            try:            
                if row[1].value is not None:
                    doc_class = 'ND'
                    try:
                        if row[12].value:
                            doc_class = str(row[12].value)[0]
                    except:

                        print('WRONG CLASS',row[0].value,row[1].value,row[12].value)
                    cat = Category(
                        #sheet_name = name,
                        code = row[0].value,
                        information = row[1].value,
                        description = row[2].value,
                        document_class = doc_class
                    )
                    session.add(cat)
                    count_cat += 1
                    session.commit()
            except:
                print('CATEGORY ERROR CODE:',row[0].value,name)
            try:
                session.commit()
            except:
                print('sheet name error', name)
    return str(count_cat) + ' Categories updated!'
    '''
    except:
        return 'Categories FAIL: check your source file.'
    '''

@rq.job('low', timeout=15)
def category_update():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    categorydel = session.query(Category).delete()
    print('deleted from Category', categorydel)
    
    category = category_upload(files['Categories'])
    print(category)

    return category

#category_update()  
def mscode_upload(source):
    print('--- - - - - - - -- - - >>>> MSCODE FUNCTION')
    wcm = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True) 
    wcm = wcm.active
    
    session = db.session
    count_mscodes = 0
    
    for row in wcm.iter_rows(min_row=2):
        print('MSCODE FUNCTION, ROW:', row)
        
        if row[0].value is not None:
            
            mscode = Mscode(

                mscode=row[0].value,
                position=int(row[1].value),
                description=row[2].value,

            )
            session.add(mscode)
            count_mscodes += 1

    session.commit()
    return str(count_mscodes) + ' MS Codes updated!'
    #except:
        #return 'MS Codes FAIL: check your source file.'

@rq.job('low', timeout=15)
def mscode_update():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    mscodedel = session.query(Mscode).delete()
    print('deleted from MSCODES', mscodedel)
    
    mscode = mscode_upload(files['MSCodes'])
    print(mscode)

    return mscode

#mscode_update() 
def init_file_type():
    '''Initialize the DB with all the File Source Type Needed'''
    session = db.session
    exist_list = [x.source_type for x in db.session.query(Sourcetype).all()]
    type_list = ['Document List','Janus','PDB','Categories','MSCodes']

    for source_type in type_list:
        if source_type not in exist_list:
            row = Sourcetype(source_type = source_type)
            session.add(row)

    session.commit()
    #flash('File Type Initialized', category='message')
  
    

    return

#init_file_type()

def check_pdb_not_in_janus():
    print('Check PDB Vs Janus')
    session = db.session
    #document_list = session.query(Doc_list).all()
    janus_list = session.query(Janus).all()
    pdb_list = session.query(Pdb).all()
    janus_ref = [x.doc_reference for x in janus_list]
    pdb_not_in_janus = []
    for doc in pdb_list:
        #print(doc.client_reference)
        if doc.doc_reference not in janus_ref:
            print('pdb not in janus', doc.client_reference, doc.revision_number, doc.doc_reference)
            pdb_not_in_janus.append(doc)
        return pdb_not_in_janus


### MDI Style Helpers
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, first_cell=None):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = first_cell
    if alignment:
        
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

### MDI Main Document Index to Excel
def mdi_excel():
    session = db.session
    # Open the MDI Template
    MDI_template = open('xls/template/MDR_Template.xlsx', mode='rb')
    wmb = load_workbook(MDI_template, data_only=True)
    ws = wmb.active

    # For every Category on each document
    categorie_list = session.query(Category).filter(Category.code != None, Category.information != None).all()
    mscodes_list = session.query(Mscode).order_by(Mscode.position).all()
    
    # Start Row for MDI - Skip Header
    start_row = 11
    end_row = 19

    fake_label = ['Purpose**','Rev.','Issue Plan','Revised Plan','Issue Actual', 'Transmittal no.', 'Return Date', 'Owner Cmmt*']

    for category in categorie_list:
        cat_code = 'Category Code: ' + category.code + ' ' + category.information
        
        # Category Section
        _ = ws.cell(row=start_row, column=1, value=cat_code)
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=5)
        _.font = Font(b=True)
        _.fill = PatternFill("solid", fgColor="DDDDDD")
        _class = ws.cell(row=start_row, column=6)

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        single = Side(border_style="medium", color="ff0000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        font = Font(b=True, color="000000")
        al = Alignment(horizontal="left", vertical="center")
        
        cat_ranges = 'A'+ str(start_row)+':E'+str(start_row)
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        style_range(ws, cat_ranges, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row,1))
        _class.fill = fill
        _class.border = border
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        #style_range(ws, org_range, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row+1,2))
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        al = Alignment(horizontal="left", vertical="center")
        v_al = Alignment(horizontal="center", vertical="center")
        
        # Style The Category Header
        x = 7
        for n in range(13):
            n = ws.cell(row=start_row, column=x)
            n.border = border
            n.fill = fill
            x += 1
        
        #### Document List Section

        document_list = session.query(Doc_list).filter(Doc_list.cat == category.code, Doc_list.mdi == True).all()
        #document_list = full_mdi(category.code)
        if document_list:      
            #print(cat_code)
            for document in document_list:
                print('-----------------',document.client_reference, document.org, document.title, category.code)
                org = ''
                if document.org:
                    org =  document.org #value['org']
                document_no = document.client_reference
                document_name = document.title
                
                classification= ''
                if document.cat_class:
                    classification = 'Class '+  document.cat_class

                #print(document)

                # Set Document Info
                
                ws.cell(row=start_row+1, column=2, value=org)
                ws.cell(row=start_row+1, column=3, value=document_no)
                ws.cell(row=start_row+1, column=4, value=document_name)
                ws.cell(row=start_row+1, column=6, value=classification)
                
                #ws.merge_cells(start_column=2, end_column=2, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=3, end_column=3, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=4, end_column=4, start_row=start_row+1, end_row=start_row+8)

                # Set Data Label
                
                
                tmp_row = start_row + 1

                for label in fake_label:    
                    _ = ws.cell(row=tmp_row, column=7, value=label)
                    dotted = Side(border_style="dotted", color="000000")
                    _.border = Border(bottom=dotted)
                    if label == "Purpose**":
                        _.border = Border(top=thin, bottom=dotted)
                    elif label == 'Owner Cmmt*':
                        _.border = Border(bottom=thin)

                    
                    tmp_row += 1
                
                
                #### PDB Section
                
                #pdb_document_list = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference).all()
                '''
                mscode_actions = set([x.required_action for x in pdb_document_list])
                if mscodes_list:
                    
                    for ms in mscode_actions:
                '''
                pdb_document = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference ).order_by(Pdb.revision_number).all()
                #janus_document = session.query(Janus).filter(Janus.client_reference_id == document.client_reference, Janus.mscode == ms ).first()
                
                # Set Janus Date on the last revision of PDB doc.
                #if janus_document:
                    
                tmp_col = 8
                tmp_row = start_row + 1

                for doc in pdb_document:
                    
                    ws.cell(row=start_row+1, column=3, value=doc.client_reference_id)
                    ws.cell(row=start_row+1, column=4, value=document.title)
                    
                    purpose = ws.cell(row=tmp_row, column=tmp_col, value=doc.document_revision_object)
                    rev = ws.cell(row=tmp_row+1, column=tmp_col, value=doc.revision_number)
                    
                    issue_plan = ''
                    revised_plan = ''
                    janus_document = session.query(Janus).filter(Janus.client_reference_id == document.client_reference, Janus.mscode == doc.required_action ).first()
                    
                    if janus_document:
                        issue_plan = janus_document.planned_date
                        revised_plan = janus_document.revised_plan_date
                    
                    issue_plan = ws.cell(row=tmp_row+2, column=tmp_col, value=issue_plan)
                    revised_plan = ws.cell(row=tmp_row+3, column=tmp_col, value=revised_plan)

                    issue_actual = ws.cell(row=tmp_row+4, column=tmp_col, value=doc.transmittal_date)
                    trans = ws.cell(row=tmp_row+5, column=tmp_col, value=doc.client_transmittal_ref_number)
                    return_date = ws.cell(row=tmp_row+6, column=tmp_col, value=doc.actual_response_date)
                    
                    #print('conversion to string***********')
                    if doc.document_status is not None:
                        owner_cmmt = str(doc.document_status)[:2]
                        print('owner***********')
                    else:
                        print('No Owner ****/////////')
                    
                    if owner_cmmt == "Tr": owner_cmmt = ""
                    
                    #owner_cmmt = 'xx'
                    owner_cmmt = ws.cell(row=tmp_row+7, column=tmp_col, value=owner_cmmt)

                    tmp_col += 1
                    print('/////here')
                    dotted = Side(border_style="dotted", color="000000")
                    border = Border(bottom=dotted, right=thin)
                    
                    al = Alignment(horizontal='center')
                    purpose.border = Border(bottom=dotted,right=thin,top=thin)
                    purpose.alignment = al
                    rev.border = border
                    rev.alignment = al
                    issue_plan.border = border
                    issue_plan.alignment = al
                    issue_plan.number_format = 'DD/MM/YYYY'
                    revised_plan.border = border
                    revised_plan.alignment = al
                    revised_plan.number_format = 'DD/MM/YYYY'
                    issue_actual.border = border
                    issue_actual.alignment = al
                    issue_actual.number_format = 'DD/MM/YYYY'
                    trans.border = border
                    trans.alignment = al
                    return_date.border = border
                    return_date.alignment = al
                    return_date.number_format = 'DD/MM/YYYY'
                    owner_cmmt.border = Border(bottom=thin,right=thin)
                    owner_cmmt.alignment = al
            
                thin = Side(border_style="thin", color="000000")
                border = Border(top=thin, left=thin, right=thin, bottom=thin)
                fill = PatternFill("solid", fgColor="DDDDDD")
                al = Alignment(horizontal="left", vertical="center")
                v_al = Alignment(horizontal="center", vertical="center")
                
                item_range = 'A'+ str(tmp_row)+':A'+str(tmp_row+7)
                org_range = 'B'+ str(tmp_row)+':B'+str(tmp_row+7)
                doc_no_range = 'C'+ str(tmp_row)+':C'+str(tmp_row+7)
                doc_name_range = 'D'+ str(tmp_row)+':E'+str(tmp_row+7)
                doc_class_range = 'F'+ str(tmp_row)+':F'+str(tmp_row+7)
                
                '''
                style_range(ws, item_range, border=border, alignment=al,first_cell=ws.cell(start_row+1,1))
                style_range(ws, org_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,2))
                style_range(ws, doc_no_range, border=border, alignment=v_al, first_cell=ws.cell(start_row+1,3))
                style_range(ws, doc_name_range, border=border, alignment=al, first_cell=ws.cell(start_row+1,4))
                style_range(ws, doc_class_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,6))
                '''
                
                #print(tmp_row)
                
                start_row += 8
        
        start_row += 1

    #print('Janus NOT Found List')
    #print(janus_not_found)
    wmb.save('xls/MDI_TEST.xlsx')

from flask_appbuilder.filemanager import uuid
#mdi_excel() 
#@rq.job('low', timeout=15)
def mdi_FULL_excel():
    ''' Include all Janus Milestone '''
    session = db.session
    # Open the MDI Template
    MDI_template = open('xls/template/MDR_Template.xlsx', mode='rb')
    wmb = load_workbook(MDI_template, data_only=True)
    ws = wmb.active

    title_by_pdb()
    update_doc_class()

    # For every Category on each document
    categorie_list = session.query(Category).filter(
        Category.code != None, 
        Category.information != None,
        ).order_by(Category.code).all()
    mscodes_list = session.query(Mscode).order_by(Mscode.position).all()
    
    # Start Row for MDI - Skip Header
    start_row = 11
    end_row = 19

    fake_label = ['Purpose**','Rev.','Issue Plan','Revised Plan','Issue Actual', 'Transmittal no.', 'Return Date', 'Owner Cmmt*']

    for category in categorie_list:
        cat_code = 'Category Code: ' + category.code + ' ' + category.information
        print(category.code, category.information)
        # Category Section
        _ = ws.cell(row=start_row, column=1, value=cat_code)
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=5)
        _.font = Font(b=True)
        _.fill = PatternFill("solid", fgColor="DDDDDD")
        _class = ws.cell(row=start_row, column=6)

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        single = Side(border_style="medium", color="ff0000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        font = Font(b=True, color="000000")
        al = Alignment(horizontal="left", vertical="center")
        
        cat_ranges = 'A'+ str(start_row)+':E'+str(start_row)
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        style_range(ws, cat_ranges, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row,1))
        _class.fill = fill
        _class.border = border
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        #style_range(ws, org_range, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row+1,2))
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        al = Alignment(horizontal="left", vertical="center")
        v_al = Alignment(horizontal="center", vertical="center")
        
        # Style The Category Header
        x = 7
        for n in range(13):
            n = ws.cell(row=start_row, column=x)
            n.border = border
            n.fill = fill
            x += 1
        
        #### Document List Section

        document_list = session.query(Doc_list).filter(Doc_list.cat == category.code).all()
        
        if document_list:      
            #print('document list len:', len(document_list))
            
            
            for document in document_list:
                #print('-----------------',document.client_reference, document.org, document.title, category.code)
                org = ''
                if document.org:
                    org =  document.org #value['org']
                document_no = document.client_reference
                document_name = document.title
                pdb_document = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference ).first()
                classification= ''
                if pdb_document.document_class:
                    classification = 'Class '+  pdb_document.document_class
                else:
                    classification = 'Class '+  document.cat_class


                #print(document)

                # Set Document Info
                
                ws.cell(row=start_row+1, column=2, value=org)
                ws.cell(row=start_row+1, column=3, value=document_no)
                ws.cell(row=start_row+1, column=4, value=document_name)
                ws.cell(row=start_row+1, column=6, value=classification)
                
                #ws.merge_cells(start_column=2, end_column=2, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=3, end_column=3, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=4, end_column=4, start_row=start_row+1, end_row=start_row+8)

                # Set Data Label
                
                
                tmp_row = start_row + 1

                for label in fake_label:    
                    _ = ws.cell(row=tmp_row, column=7, value=label)
                    dotted = Side(border_style="dotted", color="000000")
                    _.border = Border(bottom=dotted)
                    if label == "Purpose**":
                        _.border = Border(top=thin, bottom=dotted)
                    elif label == 'Owner Cmmt*':
                        _.border = Border(bottom=thin)

                    
                    tmp_row += 1
                
                
                #### PDB Section
                
                #pdb_document_list = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference).all()
                
                #pdb_document = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference ).order_by(Pdb.revision_number).all()
                janus_document = session.query(Janus).filter(Janus.client_reference_id == document.client_reference).all()
                ### Order PDB revision by letters then number | fake solution... Z0, Z1 etc..
                
                # Set Janus Date on the last revision of PDB doc.
                #if janus_document:
                #print('Before full mdi call -------******')
                pdb_document = fulmdi3(document.client_reference) 
                #print('After full mdi call -------******') 
                tmp_col = 8
                tmp_row = start_row + 1
                 
                if pdb_document: 
                    first_match = True
                    issue_type_for = ''
                    for x, doc in pdb_document:
                        if doc.client_reference_id == 'OL1-2C13-251000':
                            print(doc.client_reference_id)
                        # Issue and Revised Date only on the first janus match
                        
                        if doc is not None:
                            
                            #print('Doc in Pdb', doc.client_reference_id)
                            
                            ws.cell(row=start_row+1, column=3, value=doc.client_reference_id)
                            ws.cell(row=start_row+1, column=4, value=document.title)
                            
                            purpose = ws.cell(row=tmp_row, column=tmp_col, value=doc.document_revision_object)
                            # Clean the numeric revisions before write 
                            if doc.revision_number and doc.revision_number[0] == 'Z':
                                doc.revision_number = doc.revision_number[1:]
                            rev = ws.cell(row=tmp_row+1, column=tmp_col, value=doc.revision_number)
                            
                            issue_plan = ''
                            revised_plan = ''
                            #print('HEEEERE', document.client_reference, doc.document_revision_object)
                            janus_document = session.query(Janus).filter(
                                                            Janus.client_reference_id == document.client_reference, 
                                                            Janus.pdb_issue == doc.document_revision_object, 
                                                            #Janus.pdb_id == None
                                                            ).first()
                            if janus_document:
                                 revised_plan = janus_document.revised_plan_date
                            if doc.document_revision_object != issue_type_for:
                                issue_type_for = doc.document_revision_object
                                first_match = True
                            if janus_document and first_match:
                                #janus_document.pdb_id = pdb_document.id
                                issue_plan = janus_document.planned_date
                                #revised_plan = janus_document.revised_plan_date
                                revised_plan = doc.transmittal_date
                                pdb_doc = session.query(Pdb).filter(Pdb.client_reference_id == document.client_reference).first()
                                #print('NOT HEEEERE')
                                
                                # Moved below
                                #first_match = False
                                #if pdb_doc:
                                #    janus_document.pdb_id = pdb_doc.id
                            # If Issue Actual is there take it instead the revised plan
                            if doc.transmittal_date and first_match:
                                revised_plan = doc.transmittal_date
                            

                            if first_match == True and doc.transmittal_date is None:
                                #revised_plan = janus_document.revised_plan_date or doc.transmittal_date
                                try:
                                    revised_plan = janus_document.revised_plan_date
                                except:
                                    print('Revised Plan Date in None for', doc.client_reference_id)

                            first_match = False

                            issue_plan = ws.cell(row=tmp_row+2, column=tmp_col, value=issue_plan)
                            revised_plan = ws.cell(row=tmp_row+3, column=tmp_col, value=revised_plan)

                            issue_actual = ws.cell(row=tmp_row+4, column=tmp_col, value=doc.transmittal_date)
                            trans = ws.cell(row=tmp_row+5, column=tmp_col, value=doc.client_transmittal_ref_number)
                            return_date = ws.cell(row=tmp_row+6, column=tmp_col, value=doc.actual_response_date)
                            
                            # Owner Cmmt field
                            owner_cmmt = ''
                            if doc.document_status is not None:
                                owner_cmmt = str(doc.document_status)[:2]
                            if owner_cmmt == "Tr": owner_cmmt = ""
                            owner_cmmt = ws.cell(row=tmp_row+7, column=tmp_col, value=owner_cmmt)

                            
                            tmp_col += 1
                            
                            dotted = Side(border_style="dotted", color="000000")
                            border = Border(bottom=dotted, right=thin)
                            
                            al = Alignment(horizontal='center')
                            purpose.border = Border(bottom=dotted,right=thin,top=thin)
                            purpose.alignment = al
                            rev.border = border
                            rev.alignment = al
                            issue_plan.border = border
                            issue_plan.alignment = al
                            issue_plan.number_format = 'DD/MM/YYYY'
                            revised_plan.border = border
                            revised_plan.alignment = al
                            revised_plan.number_format = 'DD/MM/YYYY'
                            issue_actual.border = border
                            issue_actual.alignment = al
                            issue_actual.number_format = 'DD/MM/YYYY'
                            trans.border = border
                            trans.alignment = al
                            return_date.border = border
                            return_date.alignment = al
                            return_date.number_format = 'DD/MM/YYYY'
                            owner_cmmt.border = Border(bottom=thin,right=thin)
                            owner_cmmt.alignment = al
                    
                    thin = Side(border_style="thin", color="000000")
                    border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    fill = PatternFill("solid", fgColor="DDDDDD")
                    al = Alignment(horizontal="left", vertical="center")
                    v_al = Alignment(horizontal="center", vertical="center")
                    
                    item_range = 'A'+ str(tmp_row)+':A'+str(tmp_row+7)
                    org_range = 'B'+ str(tmp_row)+':B'+str(tmp_row+7)
                    doc_no_range = 'C'+ str(tmp_row)+':C'+str(tmp_row+7)
                    doc_name_range = 'D'+ str(tmp_row)+':E'+str(tmp_row+7)
                    doc_class_range = 'F'+ str(tmp_row)+':F'+str(tmp_row+7)
                    
                    '''
                    style_range(ws, item_range, border=border, alignment=al,first_cell=ws.cell(start_row+1,1))
                    style_range(ws, org_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,2))
                    style_range(ws, doc_no_range, border=border, alignment=v_al, first_cell=ws.cell(start_row+1,3))
                    style_range(ws, doc_name_range, border=border, alignment=al, first_cell=ws.cell(start_row+1,4))
                    style_range(ws, doc_class_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,6))
                    '''
                    
                    #print(tmp_row)
                    
                    start_row += 8
            
        start_row += 1

    #print('Janus NOT Found List')
    #print(janus_not_found)
    file_path = str(uuid.uuid4()) + '__sep__'+ 'MDI_TEST.xlsx' 
    wmb.save(UPLOAD_FOLDER + file_path)  
    return file_path

from .models import Mdi 
from flask_appbuilder.filemanager import uuid_namegen

@rq.job('low', timeout=3600)
def new_MDI():
    print(' ------  NEW MDI  ------ ')
    session = db.session
    file_path = mdi_FULL_excel()
    #new_file = file_path
    #new_file = '50e429ad-f9d2-4e3a-9829-15d51986bed9__sep__MDI_TEST.xlsx'
    if file_path: 
        new_mdi = Mdi(
            name = 'MDI | ' + str(datetime.today()),
            file = file_path,
            description = 'LongSon Main Document Index, generated on '+ str(datetime.now()) + ' .',
            created_by_fk = '1',
            changed_by_fk = '1'
        )
        session.add(new_mdi)
    session.commit()
    return "MDI is Done."
     
#new_MDI() 
    #print('document list len:', len(document_list))

from rq_scheduler import Scheduler
from rq import Queue, Worker

def mdi_rq():
    print(' ------  RQ MDI  ------ ')
    #new_MDI.schedule(timedelta(seconds=5))
    q = rq.get_queue(name='low')
    mdi = q.enqueue_call(new_MDI, timeout=3600)
    return 'MDI RQ Done'
    
    #print(mdi)
    

    
    # new_MDI.schedule(timedelta(seconds=5))
#
#
#mdi_rq()  
#
# 
#mdi_FULL_excel() 
#   
#@rq.job('low', timeout=15)
def pdb_list_upload2(source):
    session = db.session
    pdblist = openpyxl.load_workbook(UPLOAD_FOLDER + source, data_only=True, read_only=True)
    pdblist_ws = pdblist.active
 
    doc_list = [x.client_reference for x in session.query(Doc_list).all()]
    count_pdb = 0
    row_fail = 1         
    for row in pdblist_ws.iter_rows(min_row=2):
        row_fail += 1
        pdb = Pdb(
            doc_reference=row[0].value,
            title=row[1].value,
            revision_number=row[2].value,
            revision_date=date_parse(row[3].value),
            document_revision_object=row[4].value,

            discipline=row[6].value,
            transmittal_date=date_parse(row[16].value),
            transmittal_reference=row[8].value,
            specific_transmittal_number=row[9].value,
            required_action=row[10].value,
            response_due_date=date_parse(row[7].value),#usata per la nuova client trasmittal date
            actual_response_date=date_parse(row[12].value),
            document_status=row[13].value,
            client_transmittal_ref_number=row[14].value,
            remarks=row[15].value,
            document_class=row[17].value
        )
        pdb.created_by_fk = '1'
        
        if row[5].value:
            pdb.client_reference_id = row[5].value
        else:
            pdb.client_reference_id = 'NO-CLIENT-REFERENCES'
        
        print(pdb.client_reference_id)

        
        # Check if the pdb doc is or not in document list
        
        if row[5].value in doc_list:
            session.add(pdb)
            count_pdb += 1
        else:
            pdb.ex_client_reference = pdb.client_reference_id
            pdb.note = pdb.client_reference_id + ' | No reference in Document List.'
            pdb.client_reference_id = None
            
            
            pdb_not_in_document_list.append(pdb)

            '''
            pdb_remark = session.query(Pdb).filter(Pdb.client_reference_id == row[18].value).first()
            if pdb.client_reference_id:
                pdb_remark.client_reference_id = pdb.client_reference_id
            else:
                pdb_remark.client_reference_id = pdb.ex_client_reference
            
            '''
            '''
            doc_l = session.query(Doc_list).filter(Doc_list.client_reference == pdb.client_reference_id ).first()
            
            if doc_l is None:
                doc_l = Doc_list(
                    client_reference = pdb.client_reference_id,
                    doc_reference = pdb.doc_reference,
                    title = pdb.title
                )
                doc_l.created_by_fk = '1'
            doc_l.client_reference = pdb.client_reference_id
            doc_l.changed_by_fk = '1'
            session.add(doc_l)
        
            print('Problema con questo Remarks', pdb.client_reference_id)
            '''
        
        session.add(pdb)
        try:
            session.commit()
            
        except:
            session.remove()
            ft = session.query(Sourcetype).filter(Sourcetype.source_type == source).first()
            fs = session.query(SourceFiles).filter(SourceFiles.source_type_id == ft.id).first() 
            fs.description = 'FAIL on row ' + str(row_fail)+ ' - check your doc -> ' + row[5].value
            fs.changed_by_fk = '1'    
            session.commit()
            return 'PDB FAIL: check your source file.'
            
      
    

@rq.job('low', timeout=3600)
def pdb_update():
    session = db.session
    sf = SourceFiles
    files = dict([(str(x.source_type), x.file_source) for x in session.query(sf).all()])
    pdbdel = session.query(Pdb).delete()
    #print('deleted from PDB', pdbdel)
    pdb = pdb_list_upload2(files['PDB'])
    
    return print(pdb)

#wk = rq.get_worker('PDB UPDATE JOB') 
  

 
#pdb_update.queue() 
#pdb_update.get_worker()
#pdb_update()         
'''         
if pdb_document:
    print('PDB:',pdb_document.client_reference,pdb_document.required_action)
if janus_document:
    print('JANUS:',janus_document.client_reference,janus_document.mscode)
'''
''' 
for milestone in set([x.required_action for x in pdb_document_list] + [x.mscode for x in janus_list]):
    if milestone == code.mscode:    
        print('pdb + janus milestones:', document, milestone)    
        if pdb_document_list:
            for doc in pdb_document_list:
                print(cat_code, document, doc.title)
                if janus_list:
                    
                    for milestone in janus_list[1:]:
                        print(milestone.mscode) 
'''                        
#mdi_excel() 
# To Do: Check if PDB has 

def update_all():
    init_file_type()
    session = db.session
    
    deleted_janus = session.query(Janus).delete()
    deleted_pdb = session.query(Pdb).delete()
    
    #deleted_cat = session.query(Category).delete()
    deleted_doc = session.query(Doc_list).delete()
    
    session.commit()

    sf = SourceFiles
    files = dict([(str(x.source_type),x.file_source) for x in session.query(sf).all()])

    doc = document_list_upload(files['Document List'])
    print('doc list done')
    pdb = pdb_list_upload2(files['PDB'])
    print('pdb done')
    janus = janus_upload(files['Janus'])
    print('janus done')
    #cat = category_upload(files['Categories'])
    print('cat done')
    #milestons = mscode_update()
    

    
    #return doc, pdb, janus, cat, deleted_doc, deleted_pdb, deleted_janus, deleted_cat
    # No Category
    return doc, pdb, janus, deleted_doc, deleted_pdb, deleted_janus

#update_all() 
from .models import Category

@rq.job('low', timeout=3600)
def update_doc_class():

    session = db.session
    doc_list = session.query(Doc_list).all()
    for document in doc_list:
        try:
            cat_class = session.query(Category).filter(
                Category.code == document.cat
            ).first()
            
            document.cat_class = cat_class.document_class
            document.changed_by_fk = '1'
            

        except:
            print('Category Class Not Found for: ', document)
    session.commit()

#update_doc_class()



@rq.job('low')
def add(args):
    for n in range(1,10):
        print( '°°°°°°°°°°°°°°°° JOB IN QUEU',args,n)
    #flash('Job in queue' + str(x+y), category='message')
    return args

@rq.exception_handler
def send_alert_to_ops(job, *exc_info):
    # call other code to send alert to OPs team
    print('RQ ERROR',job, exc_info)



def test_rq(args):
    job = add.schedule(timedelta(seconds=20), args)
    #job2 = pdb_update.schedule(timedelta(seconds=10))
    queue = rq.get_queue()
    scheduler = rq.get_scheduler()
    jobs = scheduler.get_jobs()
    worker = rq.get_worker()
    
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    print('PRINT QUEUE ------------     ---------      ------------    ------ ')
    #flash('QUEUE MESSAGE',category='info')
    
    print('job:',jobs, 'queue:', queue, 'scheduler:', scheduler,'worker:', worker)
          
    return queue


#test_rq('some args')   

@rq.job('low')
def message(text, self):
    flash(text,category='info')

def fire_msg(self,text):
    job3 = message.schedule(timedelta(seconds=15), text, self)




def update_rq(self,id, source_type):

    print('UPDATE By REDIS-SERVER ------------------------- ')
    #session = db.session
    q = rq.get_queue(name='low')
    
    file_func = {
                'Document List' : document_list_update,
                'PDB' : pdb_update,
                'Janus' : janus_update,
                'Categories' : category_update,
                }
    print(file_func[str(source_type)])
    
    job = q.enqueue_call(file_func[str(source_type)], timeout=3600)
    
    
    flash(str(source_type) + ' updating...',category='info')
    return job.id

def remarks():
    session = db.session
    files = dict([(str(x.source_type),x.file_source) for x in session.query(SourceFiles).all()])

    source = files['PDB']
    pdblist = openpyxl.load_workbook(UPLOAD_FOLDER + source)
    pdblist_ws = pdblist.active
 
    
    count_pdb = 0
    remarks = []    
    for row in pdblist_ws.iter_rows(min_row=2):
        try:
            remarks.append(row[18].value)
            #print(row[5].value, '->', row[18].value)
            ex_docs = session.query(Pdb).filter(
                Pdb.client_reference_id == row[18].value
            ).all()
            for doc in ex_docs:
                print(doc.client_reference_id, row[5].value)
                doc.client_reference_id = row[5].value
        except:
            continue
            #print(row[5].value, '->', "remarks not found")
        
        session.commit()


#remarks()    

def title_by_pdb():
    session = db.session
    doc_list = session.query(Doc_list).all()
    for doc in doc_list:
        pdb = session.query(Pdb).filter(
            Pdb.client_reference_id == doc.client_reference
        ).order_by(Pdb.transmittal_date.desc()).first()
        try:
            doc.title = pdb.title
            doc.changed_by_fk = '1'
        except:
            continue
    session.commit()

#title_by_pdb()  
#mdi_rq()