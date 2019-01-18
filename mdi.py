from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, PatternFill, Color, Side, Alignment, NamedStyle
from collections import OrderedDict

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, first_cell=None):
    """
    Apply styles to a range of cells as if they were a single cell.
    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = first_cell
    if alignment:
        
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

# Open the Category Code List
CAT_List = open('xls/lists/Category_List.xlsx', mode='rb')
wcb = load_workbook(CAT_List, guess_types=True, data_only=True)

sheet_list = wcb.sheetnames
print('Sheets in Category Code List')

cat_description = {}
cat_duplicate =[]
cat_count = 0
for name in sheet_list[5:]:
    wcs = wcb[name]
    #print(name)
    for row in wcs.iter_rows(min_row=9):
        if row[1].value is not None: cat_count += 1
        if row[1].value is not None:
            print(name ,row[0].value, row[1].value)
            if row[0].value not in cat_description:
                cat_description[row[0].value] = row[1].value
            else: 
                print(row[0].value, 'is a duplicate??')
                obj = {
                    'SheetName': name,
                    'Code': row[0].value,
                    'Information': row[1].value,
                    'Description': row[2].value,
                    'Duplicate': cat_description[row[0].value]
                }
                cat_duplicate.append(obj)
    print('')        
    print('the cat count is:', cat_count)
    print('the len of category codes is:', len(cat_description))
    print('') 
    print('List of duplicate codes in category:', cat_duplicate)
    print('') 


# Opent the MDI List
MDI_list = open('xls/lists/MDI_list.xlsx', mode='rb')
wlb = load_workbook(MDI_list, guess_types=True, data_only=True)
wls = wlb['MDI_LIST']

category_list = set()
document_list = set() 

doc_index = OrderedDict()
cat_index = OrderedDict()

for row in wls.iter_rows(min_row=2):
    #print(row[0].value)
    #category.add(wls.cell(row, 1).value)

    category_list.add(row[0].value)
    document_list.add(row[2].value)
    
    obj = {
            'category_code': row[4].value,
            'title': row[3].value,
            'org': row[5].value,
            'cat_class': row[6].value,
            'class': row[7].value,
            'wght': row[8].value,
            'jan_ref': row[10].value,
            'mschain': row[11].value
            }


    # DOC INDEX
    if row[2].value in doc_index:
        doc_index[row[2].value].append(obj)
    else:
        doc_index[
            str(row[2].value)] = [obj]

    # CAT INDEX
    cat_index[str(row[0].value)] = obj

    
wjs = wlb['JANUS']
janus_list = {}

for row in wjs.iter_rows(min_row=2):
    milestone = row[2].value
    jan_ref   = row[1].value
    mscode    = row[3].value

    obj = {
        'cumulative'   : row[4].value,
        'planned_date' : row[5].value,
        'resched_date' : row[6].value
    }
    
    if milestone not in janus_list:
        janus_list[milestone] = {
            jan_ref : {
                mscode : obj
            }
        }
    elif jan_ref not in janus_list[milestone]: 
        janus_list[milestone][jan_ref] = {
            mscode : obj
        }
    elif mscode not in janus_list[milestone][jan_ref]:  
        janus_list[milestone][jan_ref][mscode] = obj


#print('Janus Pivot Objects')
#print(janus_list)
#print('Janus Test, last row mscode = START:',janus_list['E8307']['M55.F.E8307-05']['START'])

print('Document List: ', len(category_list), 'category and', len(document_list),'documents')

# Open the PDB List
PDB_list = open('xls/lists/PDB_list.xlsx', mode='rb')
wpb = load_workbook(PDB_list, guess_types=True, data_only=True)
wps = wpb.active

pdb_hash = OrderedDict()
pdb_hash_cat = {}
only_pdb_doc = False
date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

for row in wps.iter_rows(min_row=2):
    
    if row[5].value[5:8] in pdb_hash_cat:
        pdb_hash_cat[row[5].value[5:8]] += 1 
    
    else:
        pdb_hash_cat[row[5].value[5:8]] = 1

    #if row[13].value[:2] != "Tr":
    owner_cmmt = row[13].value[:2]
    if owner_cmmt == "Tr": owner_cmmt = ""
    
    if row[5].value in pdb_hash:
        #print(" ************ Already in PDB Hash")
        pdb_hash[row[5].value].append({
            'discipline': row[6].value, #ex PDB List col 15
            'document_no': row[5].value, #ex PDB List col 2
            'document_name': row[1].value, #ex PDB List col 3
            'revision': str(row[2].value), #ex PDB List col 4
            'transmittal': row[14].value, #ex PDB List col 6
            'trans_date': row[7].value, #ex PDB List col 7
            'require_action': row[4].value, #ex PDB List col 5
            'revised_plan': '',
            'issue_plan': '',
            'return_date': row[12].value, #ex PDB List col 9
            'owner_cmmt': owner_cmmt #row[13].value[:2] #ex PDB List col 3
        })
    else:
        #print(" ************ NOT in PDB Hash")
        pdb_hash[
            str(row[5].value)] = [{
            'discipline': row[6].value, #ex PDB List col 15
            'document_no': row[5].value, #ex PDB List col 2
            'document_name': row[1].value, #ex PDB List col 3
            'revision': str(row[2].value), #ex PDB List col 4
            'transmittal': row[14].value, #ex PDB List col 6
            'trans_date': row[7].value, #ex PDB List col 7
            'require_action': row[4].value, #ex PDB List col 5
            'revised_plan': '',
            'issue_plan': '',
            'return_date': row[12].value, #ex PDB List col 9
            'owner_cmmt': owner_cmmt#row[13].value[:2] #ex PDB List col 3
            }]
#print(pdb_hash)    
print('Documents in PDB', len(pdb_hash))
#print('pdb_hash_cat len', len(pdb_hash_cat))

## Check documents in PDB and NOT in MDI
pdb_not_in_mdi = []

for doc in pdb_hash:
    if doc not in document_list:
        
        pdb_not_in_mdi.append(doc)
print('')
print('Document in PDB but NOT in Document List:')

list_prev = Workbook()
list_prev_s = list_prev.active
list_previsional = []
for doc in doc_index:
    if doc not in pdb_hash:
        print(doc_index[doc][0]['org'],doc, doc_index[doc][0]['title'], doc_index[doc][0]['class'])
        row = [doc_index[doc][0]['org'],doc, doc_index[doc][0]['title'], doc_index[doc][0]['class']]

        list_previsional.append(doc)
        list_prev_s.append(row)
list_prev.save("lista_prev.xlsx")
print("prevision list len", len(list_previsional))
#print(list_previsional)

for doc in pdb_not_in_mdi: print(doc, pdb_hash[doc][0]['document_name'])

# Open the MDI Template
MDI_template = open('xls/template/MDR_Template.xlsx', mode='rb')
wmb = load_workbook(MDI_template, guess_types=True, data_only=True)
ws = wmb.active

start_row = 11
end_row = 19

fake_label = ['Purpose**','Rev.','Issue Plan','Revised Plan','Issue Actual', 'Transmittal no.', 'Return Date', 'Owner Cmmt*']
janus_not_found = []

for cat, value in cat_index.items():
    cat_code = 'Category Code: ' + cat
    if cat in cat_description:
        cat_code = 'Category Code: ' + cat + ' ' + cat_description[cat]
    item = ''
    org = value['org']
    document_no = value['title']
    document_name = str(cat)
    classification = 'Class '+ str(value['class'])

    index = cat_index
    if only_pdb_doc:
        index = pdb_hash_cat

    if cat in index:
        _ = ws.cell(row=start_row, column=1, value=cat_code)
        ws.merge_cells(start_row=start_row, end_row=start_row, start_column=1, end_column=5)
        _.font = Font(b=True)
        _.fill = PatternFill("solid", fgColor="DDDDDD")
        _class = ws.cell(row=start_row, column=6)

        thin = Side(border_style="thin", color="000000")
        double = Side(border_style="double", color="ff0000")
        single = Side(border_style="medium", color="ff0000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        font = Font(b=True, color="000000")
        al = Alignment(horizontal="left", vertical="center")
        
        cat_ranges = 'A'+ str(start_row)+':E'+str(start_row)
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        style_range(ws, cat_ranges, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row,1))
        _class.fill = fill
        _class.border = border
        #org_range = 'B'+ str(start_row+1)+':B'+str(start_row+8)
        #style_range(ws, org_range, border=border, fill=fill, font=font, alignment=al,first_cell=ws.cell(start_row+1,2))
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="DDDDDD")
        al = Alignment(horizontal="left", vertical="center")
        v_al = Alignment(horizontal="center", vertical="center")
        
        # Style The Category Header
        x = 7
        for n in range(13):
            n = ws.cell(row=start_row, column=x)
            n.border = border
            n.fill = fill
            x += 1

        
        # Set Category Code
        doc_counter = 0
        for doc, revisions in sorted(pdb_hash.items()):
            if doc[5:8] == cat:
                doc_counter += 1
                
                
                
                # Set Document Info
                
                ws.cell(row=start_row+1, column=2, value=org)
                ws.cell(row=start_row+1, column=3, value=document_no)
                ws.cell(row=start_row+1, column=4, value=document_name)
                ws.cell(row=start_row+1, column=6, value=classification)
                
                #ws.merge_cells(start_column=2, end_column=2, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=3, end_column=3, start_row=start_row+1, end_row=start_row+8)
                #ws.merge_cells(start_column=4, end_column=4, start_row=start_row+1, end_row=start_row+8)

                
                # Set Data Label
                
                
                
                tmp_row = start_row + 1

                for label in fake_label:    
                    _ = ws.cell(row=tmp_row, column=7, value=label)
                    dotted = Side(border_style="dotted", color="000000")
                    _.border = Border(bottom=dotted)
                    if label == "Purpose**":
                        _.border = Border(top=thin, bottom=dotted)
                    elif label == 'Owner Cmmt*':
                        _.border = Border(bottom=thin)

                    
                    tmp_row += 1
                
                
                    
                tmp_col = 8
                tmp_row = start_row + 1
                for revision in revisions:
                    ws.cell(row=start_row+1, column=3, value=revision['document_no'])
                    ws.cell(row=start_row+1, column=4, value=revision['document_name'])
                    
                    purpose = ws.cell(row=tmp_row, column=tmp_col, value=revision['require_action'])
                    rev = ws.cell(row=tmp_row+1, column=tmp_col, value=revision['revision'])
                    try:
                        jan_ref = doc_index[doc][0]['jan_ref']
                        milestone = doc_index[doc][0]['mschain']
                    except:
                        print('Wrong Janus reference or milestone chain for this doc', doc)
                    #print(doc,jan_ref,milestone, revision['require_action'])
                    
                    issue_plan = ''
                    revised_plan = ''
                    
                    if jan_ref and milestone:
                        #print('jan_ref and mschain found:',jan_ref, milestone)
                        try:
                            if revision['require_action'] in janus_list[milestone][jan_ref]:
                                issue_plan = janus_list[milestone][jan_ref][revision['require_action']]['planned_date']
                                revised_plan = janus_list[milestone][jan_ref][revision['require_action']]['resched_date']
                                #print(jan_ref,milestone,revision['require_action'],issue_plan, revised_plan)
                        except:
                            #print(jan_ref,milestone,revision['require_action'], 'NOT FOUND !')
                            janus_not_found.append([jan_ref,milestone,revision['require_action']])
                            
                    
                    #issue_plan = ws.cell(row=tmp_row+2, column=tmp_col, value=revision['issue_plan'])
                    
                    issue_plan = ws.cell(row=tmp_row+2, column=tmp_col, value=issue_plan)
                    revised_plan = ws.cell(row=tmp_row+3, column=tmp_col, value=revised_plan)

                    issue_actual = ws.cell(row=tmp_row+4, column=tmp_col, value=revision['trans_date'])
                    trans = ws.cell(row=tmp_row+5, column=tmp_col, value=revision['transmittal'])
                    return_date = ws.cell(row=tmp_row+6, column=tmp_col, value=revision['return_date'])
                    owner_cmmt = ws.cell(row=tmp_row+7, column=tmp_col, value=revision['owner_cmmt'])


                    tmp_col += 1

                    dotted = Side(border_style="dotted", color="000000")
                    border = Border(bottom=dotted, right=thin)
                    
                    al = Alignment(horizontal='center')
                    purpose.border = Border(bottom=dotted,right=thin,top=thin)
                    purpose.alignment = al
                    rev.border = border
                    rev.alignment = al
                    issue_plan.border = border
                    issue_plan.alignment = al
                    issue_plan.number_format = 'DD/MM/YYYY'
                    revised_plan.border = border
                    revised_plan.alignment = al
                    revised_plan.number_format = 'DD/MM/YYYY'
                    issue_actual.border = border
                    issue_actual.alignment = al
                    issue_actual.number_format = 'DD/MM/YYYY'
                    trans.border = border
                    trans.alignment = al
                    return_date.border = border
                    return_date.alignment = al
                    return_date.number_format = 'DD/MM/YYYY'
                    owner_cmmt.border = Border(bottom=thin,right=thin)
                    owner_cmmt.alignment = al
                    
                thin = Side(border_style="thin", color="000000")
                border = Border(top=thin, left=thin, right=thin, bottom=thin)
                fill = PatternFill("solid", fgColor="DDDDDD")
                al = Alignment(horizontal="left", vertical="center")
                v_al = Alignment(horizontal="center", vertical="center")
                
                item_range = 'A'+ str(tmp_row)+':A'+str(tmp_row+7)
                org_range = 'B'+ str(tmp_row)+':B'+str(tmp_row+7)
                doc_no_range = 'C'+ str(tmp_row)+':C'+str(tmp_row+7)
                doc_name_range = 'D'+ str(tmp_row)+':E'+str(tmp_row+7)
                doc_class_range = 'F'+ str(tmp_row)+':F'+str(tmp_row+7)
                
                '''
                style_range(ws, item_range, border=border, alignment=al,first_cell=ws.cell(start_row+1,1))
                style_range(ws, org_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,2))
                style_range(ws, doc_no_range, border=border, alignment=v_al, first_cell=ws.cell(start_row+1,3))
                style_range(ws, doc_name_range, border=border, alignment=al, first_cell=ws.cell(start_row+1,4))
                style_range(ws, doc_class_range, border=border, alignment=v_al,first_cell=ws.cell(start_row+1,6))
                '''
                
                print(tmp_row)
                
                start_row += 8
        
        start_row += 1

print('Janus NOT Found List')
print(janus_not_found)
wmb.save('xls/MDI_TEST.xlsx')




'''
startCol = 7
startRow = 
endCol = 
endRow = 
sheetReceiving = 
copiedData = 


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected
         

#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
def createData():
    print("Processing...")
    selectedRange = copyRange(1,2,4,14,sheet) #Change the 4 number values
    pastingRange = pasteRange(1,3,4,15,temp_sheet,selectedRange) #Change the 4 number values
    #You can save the template as another file to create a new file here too.s
    template.save("xls/MDI_TEST.xlsx")
    print("Range copied and pasted!")
'''
'''
for row in range(1,9):
    #doc_block = ws[start_row:end_row]

    ws.append(['x','y','z'])
    start_row +1

'''


